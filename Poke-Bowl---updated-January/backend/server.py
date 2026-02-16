"""
Web Server with Frame Streaming
Serves frontend and streams video + inventory data via WebSockets
"""

import asyncio
import io
import json
import logging
import sqlite3
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, Set
import base64

import cv2
import numpy as np
from aiohttp import web
import aiohttp

# Import authentication module
try:
    from auth import load_auth_config, AuthManager
    AUTH_AVAILABLE = True
except ImportError:
    AUTH_AVAILABLE = False
    logger.warning("Authentication module not available")

# Import restock manager
try:
    from restock_manager import RestockManager
    RESTOCK_AVAILABLE = True
except ImportError:
    RESTOCK_AVAILABLE = False
    logger.warning("Restock manager module not available")

logger = logging.getLogger(__name__)


class VideoStreamServer:
    """
    Lightweight async web server for streaming video and inventory data
    Uses WebSockets for low-latency bidirectional communication
    """
    
    def __init__(
        self,
        host: str = '0.0.0.0',
        port: int = 8080,
        frontend_dir: Optional[Path] = None,
        enable_auth: bool = True
    ):
        """
        Initialize web server
        
        Args:
            host: Server host address
            port: Server port
            frontend_dir: Path to frontend files
            enable_auth: Enable authentication (default: True)
        """
        self.host = host
        self.port = port
        self.frontend_dir = frontend_dir or Path(__file__).parent.parent / 'frontend'
        
        # Initialize authentication
        self.auth_enabled = False
        self.auth_manager: Optional[AuthManager] = None
        self.cookie_name = 'pb_session'
        
        if enable_auth and AUTH_AVAILABLE:
            auth_enabled, auth_manager = load_auth_config()
            if auth_enabled and auth_manager:
                self.auth_enabled = True
                self.auth_manager = auth_manager
                logger.info("Authentication enabled")
            elif auth_enabled and not auth_manager:
                # Auth enabled but not configured = deny access
                self.auth_enabled = True
                self.auth_manager = None
                logger.warning("Authentication enabled but not configured - access will be denied")
            else:
                logger.info("Authentication disabled")
        elif enable_auth and not AUTH_AVAILABLE:
            logger.warning("Authentication requested but module not available")
        
        self.app = web.Application()
        self.setup_routes()
        
        # Active WebSocket connections
        self.websockets: Set[web.WebSocketResponse] = set()
        
        # Latest frame and inventory data
        self.latest_frame: Optional[np.ndarray] = None
        self.latest_inventory: dict = {}
        self.latest_stats: dict = {}
        self.latest_freshness: dict = {}
        self.latest_sales: list = []
        self.latest_alerts: list = []
        
        # Camera management
        self._camera_ref = None          # Set via set_camera()
        self._available_cameras: list = []  # Populated at startup / on refresh
        
        # Component references for upload & analytics
        self._detector_ref = None        # Set via set_detector()
        self._inventory_tracker_ref = None  # Set via set_inventory_tracker()
        
        # Restock manager
        self.restock_manager: Optional[RestockManager] = None
        if RESTOCK_AVAILABLE:
            try:
                self.restock_manager = RestockManager()
                logger.info("Restock manager initialized")
            except Exception as e:
                logger.error(f"Failed to initialize restock manager: {e}")
        
        # Server statistics
        self.frames_streamed = 0
        self.start_time = time.time()
        
    def setup_routes(self):
        """Setup HTTP and WebSocket routes"""
        # Public routes (no authentication required)
        self.app.router.add_get('/login', self.handle_login_page)
        self.app.router.add_post('/api/login', self.handle_login)
        self.app.router.add_post('/api/logout', self.handle_logout)
        self.app.router.add_get('/health', self.handle_health)
        
        # Sub-page routes (protected) — clean URLs + .html extensions
        self.app.router.add_get('/upload', self.handle_upload_page)
        self.app.router.add_get('/upload.html', self.handle_upload_page)
        self.app.router.add_get('/analytics', self.handle_analytics_page)
        self.app.router.add_get('/analytics.html', self.handle_analytics_page)
        self.app.router.add_get('/account', self.handle_account_page)
        self.app.router.add_get('/account.html', self.handle_account_page)
        
        # Protected routes (authentication required)
        self.app.router.add_get('/', self.handle_index)
        self.app.router.add_get('/index.html', self.handle_index)
        self.app.router.add_get('/ws', self.handle_websocket)
        self.app.router.add_get('/api/stats', self.handle_stats)
        self.app.router.add_get('/api/freshness', self.handle_freshness)
        self.app.router.add_get('/api/sales', self.handle_sales)
        self.app.router.add_get('/api/alerts', self.handle_alerts)
        
        # Camera management routes
        self.app.router.add_get('/api/cameras', self.handle_list_cameras)
        self.app.router.add_post('/api/camera/switch', self.handle_switch_camera)
        
        # Upload API
        self.app.router.add_post('/api/upload/detect', self.handle_upload_detect)
        
        # Analytics API
        self.app.router.add_get('/api/analytics/summary', self.handle_analytics_summary)
        self.app.router.add_get('/api/analytics/export', self.handle_analytics_export)
        
        # Account API
        self.app.router.add_post('/api/account/change-password', self.handle_change_password)
        self.app.router.add_get('/api/account/info', self.handle_account_info)
        
        # Restock Mobile App API
        self.app.router.add_post('/api/restock/login', self.handle_restock_login)
        self.app.router.add_post('/api/restock/validate', self.handle_restock_validate)
        self.app.router.add_post('/api/restock/logout', self.handle_restock_logout)
        self.app.router.add_post('/api/restock/upload', self.handle_restock_upload)
        self.app.router.add_post('/api/restock/detect', self.handle_restock_detect)
        self.app.router.add_get('/api/restock/submissions', self.handle_restock_submissions)
        self.app.router.add_get('/api/restock/notifications', self.handle_restock_notifications)
        self.app.router.add_get('/api/restock/notifications/count', self.handle_restock_notification_count)
        self.app.router.add_post('/api/restock/notifications/read', self.handle_restock_notification_read)
        self.app.router.add_get('/api/restock/photo/{filename}', self.handle_restock_photo)
        
        # Manager API for restock moderation
        self.app.router.add_get('/api/restock/all', self.handle_restock_all)
        self.app.router.add_post('/api/restock/status', self.handle_restock_status_update)
        
    async def handle_index(self, request: web.Request) -> web.Response:
        """Serve main HTML page"""
        # Check authentication
        if not await self.check_auth(request):
            return web.HTTPFound('/login')
        
        index_path = self.frontend_dir / 'index.html'
        
        if not index_path.exists():
            return web.Response(
                text="Frontend not found. Please ensure frontend/index.html exists.",
                status=404
            )
        
        return web.FileResponse(index_path)
    
    async def handle_health(self, request: web.Request) -> web.Response:
        """Health check endpoint"""
        uptime = time.time() - self.start_time
        
        health_data = {
            'status': 'healthy',
            'uptime_seconds': uptime,
            'active_connections': len(self.websockets),
            'frames_streamed': self.frames_streamed
        }
        
        return web.json_response(health_data)
    
    async def handle_stats(self, request: web.Request) -> web.Response:
        """Return current statistics"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        return web.json_response(self.latest_stats)
    
    async def handle_freshness(self, request: web.Request) -> web.Response:
        """Return current freshness data"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        return web.json_response(self.latest_freshness)
    
    async def handle_sales(self, request: web.Request) -> web.Response:
        """Return sales log"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        limit = int(request.query.get('limit', 100))
        return web.json_response(self.latest_sales[:limit])
    
    async def handle_alerts(self, request: web.Request) -> web.Response:
        """Return alerts log"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        limit = int(request.query.get('limit', 20))
        return web.json_response(self.latest_alerts[:limit])
    
    # ------------------------------------------------------------------
    # Camera management helpers
    # ------------------------------------------------------------------

    def set_camera(self, camera):
        """
        Store a reference to the USBCamera so we can switch at runtime.

        Args:
            camera: USBCamera instance used by the streaming loop.
        """
        self._camera_ref = camera

    def set_available_cameras(self, cameras: list):
        """
        Cache the list returned by USBCamera.enumerate_cameras().

        Args:
            cameras: List of camera info dicts.
        """
        self._available_cameras = cameras

    async def handle_list_cameras(self, request: web.Request) -> web.Response:
        """
        GET /api/cameras
        Return list of available cameras and which one is active.
        Optionally pass ?refresh=1 to re-enumerate.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        # Re-enumerate if requested
        if request.query.get('refresh') == '1' and self._camera_ref is not None:
            from camera import USBCamera
            self._available_cameras = USBCamera.enumerate_cameras()

        active_index = self._camera_ref.camera_index if self._camera_ref else None

        return web.json_response({
            'cameras': self._available_cameras,
            'active_index': active_index,
        })

    async def handle_switch_camera(self, request: web.Request) -> web.Response:
        """
        POST /api/camera/switch   { "index": 2 }
        Switch the active camera to the given device index.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if self._camera_ref is None:
            return web.json_response(
                {'success': False, 'message': 'Camera not initialised'},
                status=500,
            )

        try:
            body = await request.json()
            new_index = int(body.get('index', -1))
        except Exception:
            return web.json_response(
                {'success': False, 'message': 'Invalid request body'},
                status=400,
            )

        if new_index < 0:
            return web.json_response(
                {'success': False, 'message': 'Missing or invalid camera index'},
                status=400,
            )

        # Already on this camera?
        if new_index == self._camera_ref.camera_index:
            return web.json_response({
                'success': True,
                'message': f'Already using camera {new_index}',
                'active_index': new_index,
            })

        # Perform the switch (blocks briefly while the device re-inits)
        ok = self._camera_ref.switch_camera(new_index)

        if ok:
            # Notify all WS clients
            await self._broadcast_camera_change(new_index)
            return web.json_response({
                'success': True,
                'message': f'Switched to camera {new_index}',
                'active_index': new_index,
            })
        else:
            return web.json_response(
                {'success': False, 'message': f'Failed to open camera {new_index}'},
                status=500,
            )

    async def _broadcast_camera_change(self, new_index: int):
        """Notify all WebSocket clients that the active camera changed."""
        if not self.websockets:
            return

        active_name = None
        for cam in self._available_cameras:
            if cam['index'] == new_index:
                active_name = cam['name']
                break

        message = {
            'type': 'camera_switched',
            'active_index': new_index,
            'active_name': active_name or f'Camera {new_index}',
        }

        await asyncio.gather(
            *[ws.send_json(message) for ws in self.websockets],
            return_exceptions=True,
        )

    # ------------------------------------------------------------------
    # Component setters for upload & analytics
    # ------------------------------------------------------------------

    def set_detector(self, detector):
        """Store a reference to the YOLODetector for image upload processing."""
        self._detector_ref = detector

    def set_inventory_tracker(self, tracker):
        """Store a reference to the inventory tracker for analytics queries."""
        self._inventory_tracker_ref = tracker

    # ------------------------------------------------------------------
    # Sub-page handlers
    # ------------------------------------------------------------------

    async def _serve_page(self, request: web.Request, filename: str) -> web.Response:
        """Serve a frontend HTML file with auth check."""
        if not await self.check_auth(request):
            return web.HTTPFound('/login')
        page_path = self.frontend_dir / filename
        if not page_path.exists():
            logger.error(f"Page not found: {page_path}")
            return web.Response(text=f"Page not found: {filename}", status=404)
        return web.FileResponse(page_path)

    async def handle_upload_page(self, request: web.Request) -> web.Response:
        """Serve the Upload page."""
        return await self._serve_page(request, 'upload.html')

    async def handle_analytics_page(self, request: web.Request) -> web.Response:
        """Serve the Analytics page."""
        return await self._serve_page(request, 'analytics.html')

    async def handle_account_page(self, request: web.Request) -> web.Response:
        """Serve the Account page."""
        return await self._serve_page(request, 'account.html')

    # ------------------------------------------------------------------
    # Upload API
    # ------------------------------------------------------------------

    async def handle_upload_detect(self, request: web.Request) -> web.Response:
        """
        POST /api/upload/detect
        Accept an image upload, run YOLO inference, return detections + annotated image.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if self._detector_ref is None or not self._detector_ref.is_loaded:
            return web.json_response(
                {'success': False, 'message': 'Detector not available'},
                status=503,
            )

        try:
            reader = await request.multipart()
            field = await reader.next()

            if field is None or field.name != 'image':
                return web.json_response(
                    {'success': False, 'message': 'No image field in upload'},
                    status=400,
                )

            # Read file content (limit 20 MB)
            data = await field.read(decode=False)
            if len(data) > 20 * 1024 * 1024:
                return web.json_response(
                    {'success': False, 'message': 'Image too large (max 20 MB)'},
                    status=400,
                )

            # Decode image
            nparr = np.frombuffer(data, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            if frame is None:
                return web.json_response(
                    {'success': False, 'message': 'Could not decode image'},
                    status=400,
                )

            # Run detection
            detections = self._detector_ref.detect(frame)

            # Draw detections
            annotated = self._detector_ref.draw_detections(frame, detections)

            # Encode annotated image to base64
            _, buf = cv2.imencode('.jpg', annotated, [cv2.IMWRITE_JPEG_QUALITY, 90])
            annotated_b64 = base64.b64encode(buf).decode('utf-8')

            # Summarise detections by product name
            summary = {}
            for det in detections:
                name = det['class_name']
                summary[name] = summary.get(name, 0) + 1

            return web.json_response({
                'success': True,
                'detections': detections,
                'summary': summary,
                'total_items': len(detections),
                'annotated_image': annotated_b64,
                'image_width': frame.shape[1],
                'image_height': frame.shape[0],
            })

        except Exception as e:
            logger.error(f"Upload detect error: {e}", exc_info=True)
            return web.json_response(
                {'success': False, 'message': f'Processing error: {str(e)}'},
                status=500,
            )

    # ------------------------------------------------------------------
    # Analytics API
    # ------------------------------------------------------------------

    async def handle_analytics_summary(self, request: web.Request) -> web.Response:
        """
        GET /api/analytics/summary
        Return aggregated sales and inventory analytics.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        persistence = self._get_persistence()
        if persistence is None:
            return web.json_response({
                'success': False,
                'message': 'Persistence not available',
            }, status=503)

        try:
            now = datetime.now(timezone.utc)
            now_ts = now.timestamp()

            # Time boundaries
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp()
            week_start = (now - timedelta(days=now.weekday())).replace(
                hour=0, minute=0, second=0, microsecond=0
            ).timestamp()
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp()
            thirty_days_ago = (now - timedelta(days=30)).timestamp()

            with persistence._get_connection() as conn:
                cursor = conn.cursor()

                # --- Summary counts ---
                def _sum_sales(start_ts):
                    cursor.execute(
                        "SELECT COALESCE(SUM(quantity_delta),0) FROM sales_log WHERE timestamp_utc >= ?",
                        (start_ts,),
                    )
                    return cursor.fetchone()[0]

                def _count_sales(start_ts):
                    cursor.execute(
                        "SELECT COUNT(*) FROM sales_log WHERE timestamp_utc >= ?",
                        (start_ts,),
                    )
                    return cursor.fetchone()[0]

                today_total = _sum_sales(today_start)
                today_count = _count_sales(today_start)
                week_total = _sum_sales(week_start)
                week_count = _count_sales(week_start)
                month_total = _sum_sales(month_start)
                month_count = _count_sales(month_start)

                # --- Product breakdown (last 30 days) ---
                cursor.execute("""
                    SELECT product_name,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY product_name
                    ORDER BY total_qty DESC
                """, (thirty_days_ago,))
                product_breakdown = [
                    {'product': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Daily trend (last 30 days) ---
                cursor.execute("""
                    SELECT DATE(timestamp_utc, 'unixepoch') as sale_date,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY sale_date
                    ORDER BY sale_date ASC
                """, (thirty_days_ago,))
                daily_trend = [
                    {'date': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Weekly trend (last 12 weeks) ---
                twelve_weeks_ago = (now - timedelta(weeks=12)).timestamp()
                cursor.execute("""
                    SELECT strftime('%%Y-W%%W', timestamp_utc, 'unixepoch') as sale_week,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY sale_week
                    ORDER BY sale_week ASC
                """, (twelve_weeks_ago,))
                weekly_trend = [
                    {'week': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Monthly trend (last 12 months) ---
                twelve_months_ago = (now - timedelta(days=365)).timestamp()
                cursor.execute("""
                    SELECT strftime('%%Y-%%m', timestamp_utc, 'unixepoch') as sale_month,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY sale_month
                    ORDER BY sale_month ASC
                """, (twelve_months_ago,))
                monthly_trend = [
                    {'month': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Low-stock alert frequency ---
                cursor.execute("""
                    SELECT product_name, COUNT(*) as alert_count
                    FROM alerts_log
                    WHERE alert_type = 'low_stock' AND timestamp_utc >= ?
                    GROUP BY product_name
                    ORDER BY alert_count DESC
                """, (thirty_days_ago,))
                low_stock_freq = [
                    {'product': r[0], 'alert_count': r[1]}
                    for r in cursor.fetchall()
                ]

                # --- DB stats ---
                cursor.execute("SELECT COUNT(*) FROM sales_log")
                total_sales_all = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM inventory_snapshots")
                total_snapshots = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM alerts_log")
                total_alerts = cursor.fetchone()[0]

            return web.json_response({
                'success': True,
                'summary': {
                    'today': {'total_qty': today_total, 'sale_count': today_count},
                    'week': {'total_qty': week_total, 'sale_count': week_count},
                    'month': {'total_qty': month_total, 'sale_count': month_count},
                },
                'product_breakdown': product_breakdown,
                'daily_trend': daily_trend,
                'weekly_trend': weekly_trend,
                'monthly_trend': monthly_trend,
                'low_stock_frequency': low_stock_freq,
                'database': {
                    'total_sales': total_sales_all,
                    'total_snapshots': total_snapshots,
                    'total_alerts': total_alerts,
                },
            })

        except Exception as e:
            logger.error(f"Analytics summary error: {e}", exc_info=True)
            return web.json_response(
                {'success': False, 'message': str(e)},
                status=500,
            )

    async def handle_analytics_export(self, request: web.Request) -> web.Response:
        """
        GET /api/analytics/export
        Download all historical data as an Excel (.xlsx) file.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        persistence = self._get_persistence()
        if persistence is None:
            return web.json_response(
                {'success': False, 'message': 'Persistence not available'},
                status=503,
            )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return web.json_response(
                {'success': False, 'message': 'openpyxl not installed – cannot generate Excel'},
                status=503,
            )

        try:
            wb = Workbook()

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E2936", end_color="1E2936", fill_type="solid")

            def _style_header(ws):
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

            with persistence._get_connection() as conn:
                cursor = conn.cursor()

                # --- Sales Log sheet ---
                ws_sales = wb.active
                ws_sales.title = "Sales Log"
                ws_sales.append(["ID", "Timestamp (EST)", "Product", "Qty", "Before", "After"])
                cursor.execute("SELECT id, timestamp_est, product_name, quantity_delta, inventory_before, inventory_after FROM sales_log ORDER BY timestamp_utc DESC")
                for row in cursor.fetchall():
                    ws_sales.append(list(row))
                _style_header(ws_sales)
                for col in ws_sales.columns:
                    ws_sales.column_dimensions[col[0].column_letter].width = 20

                # --- Inventory Snapshots sheet ---
                ws_inv = wb.create_sheet("Inventory Snapshots")
                ws_inv.append(["ID", "Timestamp (UTC)", "Frame#", "Total Items", "Inventory JSON"])
                cursor.execute("SELECT id, timestamp_utc, frame_number, total_items, inventory_json FROM inventory_snapshots ORDER BY timestamp_utc DESC LIMIT 5000")
                for row in cursor.fetchall():
                    r = list(row)
                    r[1] = datetime.fromtimestamp(r[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
                    ws_inv.append(r)
                _style_header(ws_inv)
                for col in ws_inv.columns:
                    ws_inv.column_dimensions[col[0].column_letter].width = 22

                # --- Alerts sheet ---
                ws_alerts = wb.create_sheet("Alerts")
                ws_alerts.append(["ID", "Timestamp (EST)", "Type", "Product", "Severity", "Message"])
                cursor.execute("SELECT id, timestamp_est, alert_type, product_name, severity, message FROM alerts_log ORDER BY timestamp_utc DESC")
                for row in cursor.fetchall():
                    ws_alerts.append(list(row))
                _style_header(ws_alerts)
                for col in ws_alerts.columns:
                    ws_alerts.column_dimensions[col[0].column_letter].width = 22

                # --- Freshness sheet ---
                ws_fresh = wb.create_sheet("Product Freshness")
                ws_fresh.append(["Product", "First Seen (UTC)", "Last Seen (UTC)", "Expired", "Expiration Days"])
                cursor.execute("SELECT product_name, first_seen_utc, last_seen_utc, is_expired, expiration_days FROM product_freshness")
                for row in cursor.fetchall():
                    r = list(row)
                    r[1] = datetime.fromtimestamp(r[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
                    r[2] = datetime.fromtimestamp(r[2], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
                    r[3] = "Yes" if r[3] else "No"
                    ws_fresh.append(r)
                _style_header(ws_fresh)
                for col in ws_fresh.columns:
                    ws_fresh.column_dimensions[col[0].column_letter].width = 24

            # Write to memory buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"veratori_data_{ts}.xlsx"

            return web.Response(
                body=buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"',
                },
            )

        except Exception as e:
            logger.error(f"Export error: {e}", exc_info=True)
            return web.json_response({'success': False, 'message': str(e)}, status=500)

    # ------------------------------------------------------------------
    # Account API
    # ------------------------------------------------------------------

    async def handle_account_info(self, request: web.Request) -> web.Response:
        """
        GET /api/account/info
        Return basic account information for the logged-in user.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        username = request.get('username', 'unknown')
        return web.json_response({
            'username': username,
            'auth_enabled': self.auth_enabled,
        })

    async def handle_change_password(self, request: web.Request) -> web.Response:
        """
        POST /api/account/change-password
        Body: { "current_password": "...", "new_password": "..." }
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if not self.auth_enabled or not self.auth_manager:
            return web.json_response(
                {'success': False, 'message': 'Authentication not configured'},
                status=503,
            )

        try:
            data = await request.json()
            current_pw = data.get('current_password', '')
            new_pw = data.get('new_password', '')

            if not current_pw or not new_pw:
                return web.json_response(
                    {'success': False, 'message': 'Both current and new passwords are required'},
                    status=400,
                )

            username = request.get('username', '')
            if not username:
                return web.json_response(
                    {'success': False, 'message': 'Session error – username not found'},
                    status=400,
                )

            success, message = self.auth_manager.change_password(username, current_pw, new_pw)
            status = 200 if success else 400
            return web.json_response({'success': success, 'message': message}, status=status)

        except Exception as e:
            logger.error(f"Change password error: {e}")
            return web.json_response(
                {'success': False, 'message': 'Server error'},
                status=500,
            )

    # ------------------------------------------------------------------
    # Helper – get persistence manager from tracker
    # ------------------------------------------------------------------

    def _get_persistence(self):
        """Return the PersistenceManager if available, else None."""
        if self._inventory_tracker_ref and hasattr(self._inventory_tracker_ref, 'persistence'):
            return self._inventory_tracker_ref.persistence
        return None

    async def handle_login_page(self, request: web.Request) -> web.Response:
        """Serve login page"""
        login_path = self.frontend_dir / 'login.html'
        
        if not login_path.exists():
            return web.Response(
                text="Login page not found.",
                status=404
            )
        
        return web.FileResponse(login_path)
    
    async def handle_login(self, request: web.Request) -> web.Response:
        """Handle login POST request"""
        if not self.auth_enabled:
            return web.json_response({'success': True, 'message': 'Authentication disabled'})
        
        if not self.auth_manager:
            return web.json_response({'success': False, 'message': 'Authentication not configured'}, status=503)
        
        try:
            data = await request.json()
            username = data.get('username', '').strip()
            password = data.get('password', '')
            
            if not username or not password:
                return web.json_response({'success': False, 'message': 'Username and password required'}, status=400)
            
            # Authenticate user
            session_token = self.auth_manager.authenticate(username, password)
            
            if session_token:
                # Create response with session cookie
                response = web.json_response({'success': True, 'message': 'Login successful'})
                
                # Determine if we're behind HTTPS
                is_secure = request.headers.get('X-Forwarded-Proto', '').lower() == 'https'
                
                # Set session cookie
                response.set_cookie(
                    self.cookie_name,
                    session_token,
                    max_age=86400,  # 24 hours
                    httponly=True,
                    samesite='Lax',
                    secure=is_secure,
                    path='/'
                )
                
                return response
            else:
                return web.json_response({'success': False, 'message': 'Invalid username or password'}, status=401)
        
        except Exception as e:
            logger.error(f"Login error: {e}")
            return web.json_response({'success': False, 'message': 'Login failed'}, status=500)
    
    async def handle_logout(self, request: web.Request) -> web.Response:
        """Handle logout POST request"""
        response = web.json_response({'success': True, 'message': 'Logged out'})
        
        # Clear session cookie
        response.del_cookie(self.cookie_name, path='/')
        
        return response
    
    async def check_auth(self, request: web.Request) -> bool:
        """
        Check if request is authenticated
        
        Args:
            request: HTTP request
            
        Returns:
            True if authenticated or auth disabled, False otherwise
        """
        if not self.auth_enabled:
            return True
        
        if not self.auth_manager:
            # Auth enabled but not configured = deny access
            return False
        
        # Get session cookie
        session_token = request.cookies.get(self.cookie_name)
        if not session_token:
            return False
        
        # Verify session
        username = self.auth_manager.verify_session(session_token)
        if username:
            # Store username in request for potential future use
            request['username'] = username
            return True
        
        return False
    
    async def handle_websocket(self, request: web.Request) -> web.WebSocketResponse:
        """
        Handle WebSocket connection for streaming
        Sends frames and inventory updates to client
        """
        # Check authentication
        if not await self.check_auth(request):
            ws = web.WebSocketResponse()
            await ws.prepare(request)
            await ws.send_json({'error': 'Unauthorized', 'type': 'error'})
            await ws.close()
            return ws
        
        ws = web.WebSocketResponse(
            heartbeat=30,  # Send ping every 30s
            compress=False  # Disable compression for lower latency
        )
        await ws.prepare(request)
        
        self.websockets.add(ws)
        client_addr = request.remote
        logger.info(f"WebSocket connected: {client_addr} (total: {len(self.websockets)})")
        
        try:
            # Send initial data
            await self.send_to_client(ws, {
                'type': 'inventory',
                'data': self.latest_inventory
            })
            
            await self.send_to_client(ws, {
                'type': 'freshness',
                'data': self.latest_freshness
            })
            
            await self.send_to_client(ws, {
                'type': 'sales',
                'data': self.latest_sales
            })
            
            await self.send_to_client(ws, {
                'type': 'alerts',
                'data': self.latest_alerts
            })
            
            # Handle incoming messages (if any)
            async for msg in ws:
                if msg.type == aiohttp.WSMsgType.TEXT:
                    try:
                        data = json.loads(msg.data)
                        await self.handle_client_message(ws, data)
                    except json.JSONDecodeError:
                        logger.warning(f"Invalid JSON from {client_addr}")
                
                elif msg.type == aiohttp.WSMsgType.ERROR:
                    logger.error(f"WebSocket error: {ws.exception()}")
        
        except asyncio.CancelledError:
            logger.info(f"WebSocket cancelled: {client_addr}")
        
        except Exception as e:
            logger.error(f"WebSocket error for {client_addr}: {e}")
        
        finally:
            self.websockets.discard(ws)
            logger.info(f"WebSocket disconnected: {client_addr} (remaining: {len(self.websockets)})")
        
        return ws
    
    async def handle_client_message(self, ws: web.WebSocketResponse, data: dict):
        """
        Handle incoming message from client
        
        Args:
            ws: WebSocket connection
            data: Parsed JSON data
        """
        msg_type = data.get('type')
        
        if msg_type == 'ping':
            await self.send_to_client(ws, {'type': 'pong'})
        
        elif msg_type == 'request_frame':
            # Client requesting latest frame
            if self.latest_frame is not None:
                await self.send_frame_to_client(ws, self.latest_frame)
        
        elif msg_type == 'switch_camera':
            # Switch camera via WebSocket
            new_index = data.get('index')
            if new_index is None or self._camera_ref is None:
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': False,
                    'message': 'Invalid index or camera not available',
                })
                return

            new_index = int(new_index)
            if new_index == self._camera_ref.camera_index:
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': True,
                    'message': f'Already using camera {new_index}',
                    'active_index': new_index,
                })
                return

            ok = self._camera_ref.switch_camera(new_index)
            if ok:
                await self._broadcast_camera_change(new_index)
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': True,
                    'active_index': new_index,
                })
            else:
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': False,
                    'message': f'Failed to open camera {new_index}',
                })
    
    async def send_to_client(self, ws: web.WebSocketResponse, data: dict):
        """
        Send JSON data to a single client
        
        Args:
            ws: WebSocket connection
            data: Data to send
        """
        try:
            await ws.send_json(data)
        except Exception as e:
            logger.error(f"Failed to send to client: {e}")
    
    async def send_frame_to_client(self, ws: web.WebSocketResponse, frame: np.ndarray):
        """
        Encode and send frame to a single client
        
        Args:
            ws: WebSocket connection
            frame: Frame to send
        """
        try:
            # Encode frame as JPEG
            _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            
            # Convert to base64
            frame_b64 = base64.b64encode(buffer).decode('utf-8')
            
            # Send with timestamp
            await ws.send_json({
                'type': 'frame',
                'data': frame_b64,
                'timestamp': time.time()
            })
        
        except Exception as e:
            logger.error(f"Failed to send frame: {e}")
    
    async def broadcast_frame(self, frame: np.ndarray):
        """
        Broadcast frame to all connected clients
        
        Args:
            frame: Frame to broadcast
        """
        if not self.websockets:
            return
        
        self.latest_frame = frame.copy()
        self.frames_streamed += 1
        
        # Encode once, send to all
        try:
            _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            frame_b64 = base64.b64encode(buffer).decode('utf-8')
            
            message = {
                'type': 'frame',
                'data': frame_b64,
                'timestamp': time.time()
            }
            
            # Send to all clients concurrently
            if self.websockets:
                await asyncio.gather(
                    *[ws.send_json(message) for ws in self.websockets],
                    return_exceptions=True
                )
        
        except Exception as e:
            logger.error(f"Failed to broadcast frame: {e}")
    
    async def broadcast_inventory(self, inventory: dict):
        """
        Broadcast inventory update to all clients
        
        Args:
            inventory: Inventory dictionary
        """
        if not self.websockets:
            self.latest_inventory = inventory
            return
        
        self.latest_inventory = inventory
        
        message = {
            'type': 'inventory',
            'data': inventory,
            'timestamp': time.time()
        }
        
        # Send to all clients
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_stats(self, stats: dict):
        """
        Broadcast statistics to all clients
        
        Args:
            stats: Statistics dictionary
        """
        self.latest_stats = stats
        
        if not self.websockets:
            return
        
        message = {
            'type': 'stats',
            'data': stats,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    def update_frame(self, frame: np.ndarray):
        """
        Update latest frame (synchronous wrapper)
        
        Args:
            frame: New frame
        """
        self.latest_frame = frame.copy()
    
    def update_inventory(self, inventory: dict):
        """
        Update latest inventory (synchronous wrapper)
        
        Args:
            inventory: New inventory
        """
        self.latest_inventory = inventory
    
    def update_stats(self, stats: dict):
        """
        Update latest stats (synchronous wrapper)
        
        Args:
            stats: New statistics
        """
        self.latest_stats = stats
    
    def update_freshness(self, freshness: dict):
        """
        Update latest freshness data (synchronous wrapper)
        
        Args:
            freshness: New freshness data
        """
        self.latest_freshness = freshness
    
    def update_sales(self, sales: list):
        """
        Update latest sales log (synchronous wrapper)
        
        Args:
            sales: New sales log entries
        """
        self.latest_sales = sales
    
    def update_alerts(self, alerts: list):
        """
        Update latest alerts (synchronous wrapper)
        
        Args:
            alerts: New alerts list
        """
        self.latest_alerts = alerts
    
    async def broadcast_freshness(self, freshness: dict):
        """
        Broadcast freshness update to all clients
        
        Args:
            freshness: Freshness dictionary
        """
        self.latest_freshness = freshness
        
        if not self.websockets:
            return
        
        message = {
            'type': 'freshness',
            'data': freshness,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_sales(self, sales: list):
        """
        Broadcast sales log update to all clients
        
        Args:
            sales: Sales log entries list
        """
        self.latest_sales = sales
        
        if not self.websockets:
            return
        
        message = {
            'type': 'sales',
            'data': sales,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_alerts(self, alerts: list):
        """
        Broadcast alerts update to all clients
        
        Args:
            alerts: Alerts list
        """
        self.latest_alerts = alerts
        
        if not self.websockets:
            return
        
        message = {
            'type': 'alerts',
            'data': alerts,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    # ------------------------------------------------------------------
    # Restock Mobile App API Handlers
    # ------------------------------------------------------------------
    
    async def handle_restock_login(self, request: web.Request) -> web.Response:
        """Handle restock app login"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Restock service unavailable'}, status=503)
        
        try:
            data = await request.json()
            username = data.get('username')
            password = data.get('password')
            
            if not username or not password:
                return web.json_response({'success': False, 'message': 'Username and password required'}, status=400)
            
            # Authenticate using existing auth system
            if not self.auth_manager:
                return web.json_response({'success': False, 'message': 'Authentication not configured'}, status=503)
            
            if not self.auth_manager.verify_password(username, password):
                return web.json_response({'success': False, 'message': 'Invalid credentials'}, status=401)
            
            # Create session token
            session_token = self.auth_manager.authenticate(username, password)
            if not session_token:
                return web.json_response({'success': False, 'message': 'Authentication failed'}, status=401)
            
            # Get user role and franchise (simplified - would come from user database)
            role = 'employee'  # Would be determined from user database
            franchise = 'f1'  # Would be determined from user database
            
            # Return session info with cookie
            response = web.json_response({
                'success': True,
                'username': username,
                'role': role,
                'franchise': franchise,
            })
            
            # Set session cookie
            is_secure = request.headers.get('X-Forwarded-Proto', '').lower() == 'https'
            response.set_cookie(
                self.cookie_name,
                session_token,
                max_age=86400,  # 24 hours
                httponly=True,
                samesite='Lax',
                secure=is_secure,
                path='/'
            )
            
            return response
        
        except Exception as e:
            logger.error(f"Restock login error: {e}")
            return web.json_response({'success': False, 'message': 'Login failed'}, status=500)
    
    async def handle_restock_validate(self, request: web.Request) -> web.Response:
        """Validate restock session token"""
        try:
            # Get session cookie
            session_token = request.cookies.get(self.cookie_name)
            if not session_token:
                return web.json_response({'valid': False}, status=401)
            
            if not self.auth_manager:
                return web.json_response({'valid': False}, status=503)
            
            username = self.auth_manager.verify_session(session_token)
            if username:
                return web.json_response({'valid': True, 'username': username})
            else:
                return web.json_response({'valid': False}, status=401)
        
        except Exception as e:
            logger.error(f"Restock validate error: {e}")
            return web.json_response({'valid': False}, status=500)
    
    async def handle_restock_logout(self, request: web.Request) -> web.Response:
        """Handle restock app logout"""
        response = web.json_response({'success': True, 'message': 'Logged out'})
        response.del_cookie(self.cookie_name, path='/')
        return response
    
    async def handle_restock_detect(self, request: web.Request) -> web.Response:
        """Run YOLO detection on uploaded photo (for preview)"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Service unavailable'}, status=503)
        
        if not self._detector_ref:
            return web.json_response({'success': False, 'message': 'Detector not available'}, status=503)
        
        try:
            # Parse multipart form data
            reader = await request.multipart()
            photo_data = None
            
            async for field in reader:
                if field.name == 'photo':
                    photo_data = await field.read()
                    break
            
            if not photo_data:
                return web.json_response({'success': False, 'message': 'No photo provided'}, status=400)
            
            # Convert to numpy array
            nparr = np.frombuffer(photo_data, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if img is None:
                return web.json_response({'success': False, 'message': 'Invalid image'}, status=400)
            
            # Run detection
            detections = self._detector_ref.detect(img)
            
            # Count products by class
            product_counts = {}
            for det in detections:
                class_name = det.get('class_name', 'unknown')
                product_counts[class_name] = product_counts.get(class_name, 0) + 1
            
            return web.json_response({
                'success': True,
                'detections': detections,
                'product_counts': product_counts,
                'total_detections': len(detections)
            })
        
        except Exception as e:
            logger.error(f"Restock detect error: {e}")
            return web.json_response({'success': False, 'message': 'Detection failed'}, status=500)
    
    async def handle_restock_upload(self, request: web.Request) -> web.Response:
        """Handle restock photo upload"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Restock service unavailable'}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'message': 'Unauthorized'}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            # Parse multipart form data
            reader = await request.multipart()
            
            photos = []
            station = None
            product = None
            notes = None
            device_id = None
            latitude = None
            longitude = None
            detection_results = None
            franchise_id = 'f1'  # Would come from user database
            
            async for field in reader:
                if field.name.startswith('photo_'):
                    photo_data = await field.read()
                    if photo_data:
                        photos.append(photo_data)
                elif field.name == 'station':
                    station = await field.text()
                elif field.name == 'product':
                    product = await field.text()
                elif field.name == 'notes':
                    notes = await field.text()
                elif field.name == 'device_id':
                    device_id = await field.text()
                elif field.name == 'latitude':
                    lat_str = await field.text()
                    latitude = float(lat_str) if lat_str else None
                elif field.name == 'longitude':
                    lng_str = await field.text()
                    longitude = float(lng_str) if lng_str else None
                elif field.name == 'detection_results':
                    detection_str = await field.text()
                    if detection_str:
                        import json
                        detection_results = json.loads(detection_str)
            
            if not station or not product:
                return web.json_response({'success': False, 'message': 'Station and product required'}, status=400)
            
            if len(photos) < 3:
                return web.json_response({'success': False, 'message': 'Minimum 3 photos required'}, status=400)
            
            # Create submission
            success, message, submission_id = self.restock_manager.create_submission(
                employee_username=username,
                franchise_id=franchise_id,
                station=station,
                product=product,
                notes=notes,
                device_id=device_id,
                latitude=latitude,
                longitude=longitude,
                photos=photos,
                detection_results=detection_results
            )
            
            if not success:
                return web.json_response({'success': False, 'message': message}, status=400)
            
            return web.json_response({
                'success': True,
                'message': message,
                'submission_id': submission_id
            })
        
        except Exception as e:
            logger.error(f"Restock upload error: {e}")
            return web.json_response({'success': False, 'message': 'Upload failed'}, status=500)
    
    async def handle_restock_submissions(self, request: web.Request) -> web.Response:
        """Get employee's submissions"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'submissions': []}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'submissions': []}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            submissions = self.restock_manager.get_employee_submissions(username)
            return web.json_response({'success': True, 'submissions': submissions})
        
        except Exception as e:
            logger.error(f"Error getting submissions: {e}")
            return web.json_response({'success': False, 'submissions': []}, status=500)
    
    async def handle_restock_notifications(self, request: web.Request) -> web.Response:
        """Get employee notifications"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'notifications': []}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'notifications': []}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            notifications = self.restock_manager.get_notifications(username)
            formatted = [{
                'id': n['id'],
                'title': n['title'],
                'message': n['message'],
                'timestamp': n['timestamp_utc'],
                'read': bool(n['read'])
            } for n in notifications]
            
            return web.json_response({'success': True, 'notifications': formatted})
        
        except Exception as e:
            logger.error(f"Error getting notifications: {e}")
            return web.json_response({'success': False, 'notifications': []}, status=500)
    
    async def handle_restock_notification_count(self, request: web.Request) -> web.Response:
        """Get unread notification count"""
        if not self.restock_manager:
            return web.json_response({'count': 0}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'count': 0}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            count = self.restock_manager.get_notification_count(username)
            return web.json_response({'count': count})
        except Exception as e:
            logger.error(f"Error getting notification count: {e}")
            return web.json_response({'count': 0}, status=500)
    
    async def handle_restock_notification_read(self, request: web.Request) -> web.Response:
        """Mark notification as read"""
        if not self.restock_manager:
            return web.json_response({'success': False}, status=503)
        
        try:
            data = await request.json()
            notification_id = data.get('notification_id')
            
            if not notification_id:
                return web.json_response({'success': False}, status=400)
            
            success = self.restock_manager.mark_notification_read(str(notification_id))
            return web.json_response({'success': success})
        except Exception as e:
            logger.error(f"Error marking notification read: {e}")
            return web.json_response({'success': False}, status=500)
    
    async def handle_restock_photo(self, request: web.Request) -> web.Response:
        """Serve restock photo file"""
        if not self.restock_manager:
            return web.Response(status=404)
        
        filename = request.match_info.get('filename')
        if not filename:
            return web.Response(status=404)
        
        photo_path = self.restock_manager.get_photo_path(filename)
        if photo_path and photo_path.exists():
            return web.FileResponse(photo_path)
        
        return web.Response(status=404)
    
    async def handle_restock_all(self, request: web.Request) -> web.Response:
        """Get all submissions (manager only)"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'submissions': []}, status=503)
        
        # Check authentication and manager role
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'submissions': []}, status=401)
        
        try:
            # Get query parameters
            franchise_id = request.query.get('franchise')
            status = request.query.get('status')
            employee = request.query.get('employee')
            
            submissions = self.restock_manager.get_all_submissions(
                franchise_id=franchise_id,
                status=status,
                employee=employee
            )
            
            return web.json_response({'success': True, 'submissions': submissions})
        
        except Exception as e:
            logger.error(f"Error getting all submissions: {e}")
            return web.json_response({'success': False, 'submissions': []}, status=500)
    
    async def handle_restock_status_update(self, request: web.Request) -> web.Response:
        """Update submission status (manager only)"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Service unavailable'}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'message': 'Unauthorized'}, status=401)
        
        try:
            data = await request.json()
            submission_id = data.get('submission_id')
            status = data.get('status')
            feedback = data.get('feedback')
            
            if not submission_id or not status:
                return web.json_response({'success': False, 'message': 'Missing required fields'}, status=400)
            
            # Get reviewer username
            username = request.get('username', 'manager')
            
            success = self.restock_manager.update_submission_status(
                submission_id=submission_id,
                status=status,
                reviewed_by=username,
                feedback=feedback
            )
            
            if success:
                return web.json_response({'success': True, 'message': 'Status updated'})
            else:
                return web.json_response({'success': False, 'message': 'Update failed'}, status=400)
        
        except Exception as e:
            logger.error(f"Error updating submission status: {e}")
            return web.json_response({'success': False, 'message': 'Update failed'}, status=500)
    
    async def start(self):
        """Start the web server"""
        runner = web.AppRunner(self.app)
        await runner.setup()
        
        site = web.TCPSite(runner, self.host, self.port)
        await site.start()
        
        logger.info(f"Server started at http://{self.host}:{self.port}")
        logger.info(f"Frontend directory: {self.frontend_dir}")
    
    def run(self):
        """Run the server (blocking)"""
        web.run_app(
            self.app,
            host=self.host,
            port=self.port,
            print=None,  # Disable aiohttp's startup message
            access_log=None  # Disable access logs for performance
        )
    
    def get_url(self) -> str:
        """
        Get server URL
        
        Returns:
            Server URL string
        """
        return f"http://{self.host}:{self.port}"


class StreamManager:
    """
    Manages streaming loop coordination between camera, detector, and server
    """
    
    def __init__(
        self,
        camera,
        detector,
        inventory_tracker,
        server: VideoStreamServer,
        target_fps: int = 30
    ):
        """
        Initialize stream manager
        
        Args:
            camera: USBCamera instance
            detector: YOLODetector instance
            inventory_tracker: InventoryTracker instance
            server: VideoStreamServer instance
            target_fps: Target streaming FPS
        """
        self.camera = camera
        self.detector = detector
        self.inventory_tracker = inventory_tracker
        self.server = server
        self.target_fps = target_fps
        self.frame_interval = 1.0 / target_fps
        
        self.is_running = False
        self.loop_task = None
    
    async def stream_loop(self):
        """
        Main streaming loop
        Captures frames, runs inference, updates inventory, and broadcasts
        """
        logger.info("Starting stream loop...")
        self.is_running = True
        
        frame_count = 0
        last_stats_time = time.time()
        stats_interval = 1.0  # Update stats every second
        
        while self.is_running:
            loop_start = time.time()
            
            # Capture frame
            success, frame = self.camera.read()
            
            if not success or frame is None:
                logger.warning("Failed to capture frame, attempting reconnection...")
                if not self.camera.reconnect():
                    await asyncio.sleep(1.0)
                    continue
                success, frame = self.camera.read()
                if not success:
                    await asyncio.sleep(1.0)
                    continue
            
            # Run detection
            detections = self.detector.detect(frame)
            
            # Update inventory
            self.inventory_tracker.update(detections)
            inventory = self.inventory_tracker.get_inventory()
            
            # Draw detections on frame
            annotated_frame = self.detector.draw_detections(frame, detections)
            
            # Broadcast frame and inventory
            await self.server.broadcast_frame(annotated_frame)
            await self.server.broadcast_inventory(inventory)
            
            frame_count += 1
            
            # Broadcast stats periodically
            current_time = time.time()
            if current_time - last_stats_time >= stats_interval:
                stats = {
                    'fps': self.detector.get_fps(),
                    'inference_time': self.detector.get_average_inference_time(),
                    'total_items': self.inventory_tracker.get_total_items(),
                    'frame_count': frame_count,
                    'active_connections': len(self.server.websockets)
                }
                await self.server.broadcast_stats(stats)
                
                # Broadcast freshness, sales, and alerts data if available
                if hasattr(self.inventory_tracker, 'get_freshness_state'):
                    freshness = self.inventory_tracker.get_freshness_state()
                    await self.server.broadcast_freshness(freshness)
                
                if hasattr(self.inventory_tracker, 'get_sales_history'):
                    sales = self.inventory_tracker.get_sales_history(limit=100)
                    await self.server.broadcast_sales(sales)
                
                if hasattr(self.inventory_tracker, 'get_recent_alerts'):
                    alerts = self.inventory_tracker.get_recent_alerts(limit=20)
                    await self.server.broadcast_alerts(alerts)
                
                last_stats_time = current_time
            
            # Maintain target FPS
            elapsed = time.time() - loop_start
            sleep_time = max(0, self.frame_interval - elapsed)
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
        
        logger.info("Stream loop stopped")
    
    def start(self):
        """Start streaming loop"""
        if self.is_running:
            logger.warning("Stream already running")
            return
        
        self.loop_task = asyncio.create_task(self.stream_loop())
    
    async def stop(self):
        """Stop streaming loop"""
        self.is_running = False
        
        if self.loop_task:
            await self.loop_task
            self.loop_task = None

